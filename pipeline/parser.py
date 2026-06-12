"""Parse CSKH and MB_Email Excel files from bytes — no disk I/O."""
import hashlib
import io
from datetime import datetime

import openpyxl

from pipeline.classify import _norm, detect_email_product

_LOAI_MAP_NEW = {
    "cuộc gọi đến":    "Gọi vào",
    "cuộc gọi ra":     "Gọi ra",
    "cuộc gọi đi":     "Gọi ra",
    "email - litex":   "Email LiteX",
    "email":           "Email LiteX",
    "email - mb 24/7": "Email MB",
    "email mb247":     "Email MB",
    "mạng xã hội":    "Mạng xã hội",
}

_LOAI_MAP_OLD = {
    "email":          "Email LiteX",
    "email mb247":    "Email MB",
    "cuộc gọi đi":   "Gọi ra",
    "mạng xã hội":   "Mạng xã hội",
}

_KQMAP_NEW = {
    "kq_huy_dv":              "Yêu cầu hủy",
    "kq_huong_dan_huy_app":   "Yêu cầu hủy",
    "kq_tiep_tuc_su_dung":    "KH tiếp tục sử dụng",
    "kq_khong_co_kq_huy":     "Không có kết quả",
    "kq_khong_co_kq_khac":    "Không có kết quả",
    "kq_khong_co_kq_bt":      "Không có kết quả",
    "kq_hoan_thanh_tu_van":   "Hoàn thành tư vấn",
    "kq_hoan_thanh_khac":     "Hoàn thành tư vấn",
}

# Values that mean "no real name provided"
_EMPTY_NAME = {"kh", "khách hàng", "kh.", "n/a", "none", ""}


def _parse_date(val) -> "datetime.date | None":
    if val is None:
        return None
    if hasattr(val, "date"):
        return val.date()
    if isinstance(val, str):
        s = val.strip()
        for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _make_old_format_id(filename: str, row_vals: tuple) -> str:
    raw = filename + "|" + "|".join("" if v is None else str(v) for v in row_vals)
    return "old_" + hashlib.sha256(raw.encode("utf-8", errors="replace")).hexdigest()[:32]


def _get_product(raw: str, product_normalize: dict) -> str:
    key = raw.strip()
    if key in product_normalize:
        return product_normalize[key]
    if key in product_normalize.values():
        return key
    return next(iter(product_normalize.values()), "Khác")


def _clean_ten_kh(raw) -> str:
    if not raw:
        return ""
    s = _norm(str(raw).strip())
    return "" if s in _EMPTY_NAME else str(raw).strip()


def _is_new_format(ws) -> bool:
    h = ws.cell(1, 3).value or ""
    return "loại cuộc gọi" not in _norm(str(h))


def parse_cskh_bytes(
    file_bytes: bytes,
    filename: str,
    product_normalize: dict,
) -> list[dict]:
    """Parse a CSKH Excel file (bytes) into rows ready for cskh_raw upsert.

    Returns list of dicts with keys:
      id, ma_phieu, source_file, format, event_date,
      loai, loai_kn, noi_dung, ket_qua, product, thai_do, ten_kh
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    is_new = _is_new_format(ws)
    rows = []

    if is_new:
        for row_tuple in ws.iter_rows(min_row=2, values_only=True):
            if len(row_tuple) < 16:
                continue
            ma_phieu    = row_tuple[1]   # col 2
            ten_kh_raw  = row_tuple[2]   # col 3
            loai_raw    = row_tuple[7]   # col 8
            loai_kn     = row_tuple[9]   # col 10
            thai_do_raw = row_tuple[11]  # col 12
            noi_dung    = row_tuple[12]  # col 13
            sp_raw      = row_tuple[13]  # col 14
            kq_raw      = row_tuple[14]  # col 15
            d_val       = row_tuple[15]  # col 16

            dt = _parse_date(d_val)
            if dt is None or not loai_raw:
                continue

            loai    = _LOAI_MAP_NEW.get(_norm(str(loai_raw).strip()), str(loai_raw).strip())
            kq_key  = str(kq_raw).strip() if kq_raw else ""
            kq      = _KQMAP_NEW.get(kq_key, kq_key)
            product = _get_product(str(sp_raw).strip() if sp_raw else "", product_normalize)

            rows.append({
                "id":          str(ma_phieu) if ma_phieu is not None else _make_old_format_id(filename, (row_tuple,)),
                "ma_phieu":    str(ma_phieu) if ma_phieu is not None else None,
                "source_file": filename,
                "format":      "new",
                "event_date":  dt,
                "loai":        loai,
                "loai_kn":     str(loai_kn).strip() if loai_kn else "",
                "noi_dung":    str(noi_dung).strip() if noi_dung else "",
                "ket_qua":     kq,
                "product":     product,
                "thai_do":     str(thai_do_raw).strip() if thai_do_raw else "",
                "ten_kh":      _clean_ten_kh(ten_kh_raw),
            })
    else:
        for row_idx, row_tuple in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if len(row_tuple) < 16:
                continue
            d_val       = row_tuple[1]   # col 2
            loai_raw    = row_tuple[2]   # col 3
            ten_kh_raw  = row_tuple[4]   # col 5
            loai_kn     = row_tuple[10]  # col 11
            thai_do_raw = row_tuple[12]  # col 13
            noi_dung    = row_tuple[13]  # col 14
            sp_raw      = row_tuple[14]  # col 15
            ket_qua     = row_tuple[15]  # col 16

            dt = _parse_date(d_val)
            if dt is None or not loai_raw:
                continue

            # Include row index + first 16 cols — row index guarantees uniqueness when
            # two rows are identical in all 16 columns (avoids PRIMARY KEY collision).
            row_vals = (row_idx,) + row_tuple[:16]
            row_id   = _make_old_format_id(filename, row_vals)

            loai    = _LOAI_MAP_OLD.get(_norm(str(loai_raw).strip()), str(loai_raw).strip())
            product = _get_product(str(sp_raw).strip() if sp_raw else "", product_normalize)

            rows.append({
                "id":          row_id,
                "ma_phieu":    None,
                "source_file": filename,
                "format":      "old",
                "event_date":  dt,
                "loai":        loai,
                "loai_kn":     str(loai_kn).strip() if loai_kn else "",
                "noi_dung":    str(noi_dung).strip() if noi_dung else "",
                "ket_qua":     str(ket_qua).strip() if ket_qua else "",
                "product":     product,
                "thai_do":     str(thai_do_raw).strip() if thai_do_raw else "",
                "ten_kh":      _clean_ten_kh(ten_kh_raw),
            })

    wb.close()
    return rows


def parse_mb_email_bytes(
    file_bytes: bytes,
    filename: str,
    product_keywords: list[tuple[str, str]],
) -> list[dict]:
    """Parse an MB_Email Excel file (bytes) into rows ready for mb_email_raw upsert.

    Returns list of dicts with keys:
      ticket_id, source_file, event_date, content, product
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = []

    for row_tuple in ws.iter_rows(min_row=2, values_only=True):
        if len(row_tuple) < 19:
            continue
        ticket_id = row_tuple[0]   # col 1
        d_val     = row_tuple[9]   # col 10
        content   = row_tuple[18]  # col 19

        if not d_val or not ticket_id:
            continue

        dt = _parse_date(d_val)
        if dt is None:
            continue

        product = detect_email_product(str(content) if content else "", product_keywords)

        rows.append({
            "ticket_id":   str(ticket_id),
            "source_file": filename,
            "event_date":  dt,
            "content":     str(content) if content else "",
            "product":     product,
        })

    wb.close()
    return rows
