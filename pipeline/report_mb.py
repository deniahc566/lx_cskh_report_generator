"""Build the MB (reclassified) CSKH report — returns bytes (BytesIO), no disk I/O."""
from __future__ import annotations

import io
import sys
from collections import defaultdict
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from pipeline.calc import (
    calc_mb, VALID_DATE_RANGE, HARDCODED_METRICS_MB,
    PRODUCT_START_DATES, DATE_GROUPS, _build_columns,
)

# ── Colours ───────────────────────────────────────────────────────────────────
C_TITLE      = "1F3864"
C_SUBTITLE   = "2E75B6"
C_HEADER     = "1F4E79"
C_SECTION_NO = "1F4E79"
C_SECTION    = "BDD7EE"
C_BOLD_ROW   = "DEEAF1"
C_NORMAL     = "FFFFFF"
C_SUBITEM    = "EBF3FB"
C_LUY_KE     = "E2EFDA"
C_TY_LE      = "FCE4D6"


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=11, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)


def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border():
    thin = Side(style="thin", color="B0B0B0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _NOTE_TOTAL(lk: dict, kha: int) -> str:
    ty_le_total = lk["total"] / kha if kha else 0
    ty_le_huy   = lk["huy"] / lk["total"] if lk["total"] else 0
    return (
        f"Tỷ lệ thắc mắc/khiếu nại so với tổng số KH là {ty_le_total:.2%} so với trung bình "
        f"trên các kênh khác là 0.4%\n"
        f"Số cuộc gọi nhiều do KH chưa chủ động hủy được bảo hiểm chiếm {ty_le_huy:.1%} tổng số cuộc gọi"
    )


def _NOTE_HUY(lk: dict, kha: int) -> str:
    ty_le = lk["so_huy"] / kha if kha else 0
    return f"Tỷ lệ hủy bảo hiểm là {ty_le:.2%} so với trung bình các kênh khác là 5%"


def _NOTE_TT(lk: dict, kha: int) -> str:
    ty_le = (lk["total"] - lk["so_tiep_tuc"]) / kha if kha else 0
    return (
        f"Nếu trừ đi lượng KH tiếp tục sử dụng thì tỷ lệ thắc mắc/khiếu nại "
        f"so với tổng số KH là {ty_le:.2%}"
    )


ROWS_DEF = [
    (0, "",  "Tổng Số lượng cuộc gọi",                                              "total",       True,  False, False, _NOTE_TOTAL),
    (0, "1", "Thái độ khách hàng",                                                  None,          True,  True,  False, ""),
    (0, "",  "Hài lòng",                                                            "thai_do_hai_long",    False, False, False, ""),
    (0, "",  "Bình thường",                                                         "thai_do_binh_thuong", False, False, False, ""),
    (0, "",  "Gay gắt",                                                             "thai_do_gay_gat",     False, False, False, ""),
    (0, "2", "Kênh tiếp nhận",                                                       None,          True,  True,  False, ""),
    (0, "",  "MB",                                                                    "mb_total",    True,  False, False, ""),
    (0, "",  "Call",                                                                  "mb_call",     False, False, False, ""),
    (0, "",  "Email",                                                                 "mb_email",    False, False, False, ""),
    (0, "",  "Fanpage / mạng xã hội / website",                                      "mb_fanpage",  False, False, False, ""),
    (0, "",  "LiteX",                                                                 "litex_total", True,  False, False, ""),
    (0, "",  "Call",                                                                  "litex_call",  False, False, False, ""),
    (0, "",  "Email",                                                                 "litex_email", False, False, False, ""),
    (0, "",  "Fanpage / mạng xã hội / website",                                      "litex_fanpage", False, False, False, ""),
    (0, "3", "Phân loại cuộc gọi",                                                   None,          True,  True,  False, ""),
    (0, "",  "Khách hàng hỏi quyền lợi sản phẩm / thao tác",                        "tu_van",      False, False, False, ""),
    (0, "",  "Khách hàng yêu cầu hủy dịch vụ",                                      "huy",         False, False, False, ""),
    (1, "",  "- Số khách hàng hủy bảo hiểm",                                        "so_huy",      False, False, True,  _NOTE_HUY),
    (1, "",  "- Số khách hàng tiếp tục sử dụng",                                    "so_tiep_tuc", False, False, True,  _NOTE_TT),
    (1, "",  "- Không có kết quả (KH mất kết nối, ngắt máy, không cần hỗ trợ nữa)", "so_chua_kq", False, False, True,  ""),
    (0, "",  "Số khách hàng có khiếu nại khác",                                      "khac",        False, False, False, ""),
    (0, "",  "Khách hàng yêu cầu bồi thường",                                       "boi_thuong",  False, False, False, ""),
]


def _get_date(r: dict):
    return r.get("date") or r.get("event_date")


def _populate_sheet_mb(ws, data: list[dict], kh_active: int, product_name: str = ""):
    data = [r for r in data if VALID_DATE_RANGE[0] <= _get_date(r) <= VALID_DATE_RANGE[1]]

    product_start = PRODUCT_START_DATES.get(product_name)
    if product_start:
        data = [r for r in data if _get_date(r) >= product_start]

    dates = sorted({_get_date(r) for r in data})
    by_date: dict[date, list] = defaultdict(list)
    for r in data:
        by_date[_get_date(r)].append(r)

    metrics_by_date = {d: calc_mb(by_date[d]) for d in dates}
    if product_name == "Mất Tiền MB Đền":
        metrics_by_date.update(HARDCODED_METRICS_MB)
        all_dates = sorted(set(dates) | set(HARDCODED_METRICS_MB))
    else:
        all_dates = sorted(set(dates))

    columns = _build_columns(all_dates, metrics_by_date)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    for i in range(len(columns)):
        ws.column_dimensions[get_column_letter(i + 3)].width = 14
    luy_ke_col = len(columns) + 3
    ty_le_col  = luy_ke_col + 1
    note_col   = ty_le_col + 1
    ws.column_dimensions[get_column_letter(luy_ke_col)].width = 14
    ws.column_dimensions[get_column_letter(ty_le_col)].width  = 22
    ws.column_dimensions[get_column_letter(note_col)].width   = 55
    total_cols = note_col

    date_range_str = (
        f"{all_dates[0].strftime('%d/%m/%Y')} - {all_dates[-1].strftime('%d/%m/%Y')}"
        if len(all_dates) > 1 else all_dates[0].strftime("%d/%m/%Y")
    )
    timestamp_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    def merge_and_style(row, c1, c2, text, bg, fg="FFFFFF", bold=True, size=13, italic=False, h="center"):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row, c1, text)
        cell.fill = _fill(bg); cell.font = _font(bold=bold, color=fg, size=size, italic=italic)
        cell.alignment = _align(h=h); cell.border = _border()

    ws.row_dimensions[1].height = 36
    title = f"Báo cáo dịch vụ khách hàng - Sản phẩm {product_name}" if product_name else "Báo cáo dịch vụ khách hàng - MB"
    merge_and_style(1, 1, total_cols, title, C_TITLE, size=16)

    ws.row_dimensions[2].height = 20
    merge_and_style(2, 1, total_cols, f"Thời gian lấy dữ liệu: {date_range_str}", C_SUBTITLE, size=12)

    ws.row_dimensions[3].height = 20
    merge_and_style(3, 1, total_cols,
                    f"Số khách hàng active: {kh_active:,}    |    Cập nhật lần cuối: {timestamp_str}",
                    C_SUBTITLE, size=12, bold=False, italic=True)

    ws.row_dimensions[4].height = 30
    headers = ["STT", "Chỉ tiêu / Nội dung"] + [lbl for lbl, _ in columns] + ["Lũy kế", "Tỷ lệ so với số KH active", "Ghi chú"]
    for c, hdr in enumerate(headers, start=1):
        cell = ws.cell(4, c, hdr)
        cell.fill = _fill(C_HEADER); cell.font = _font(bold=True, color="FFFFFF", size=11)
        cell.alignment = _align(); cell.border = _border()

    luy_ke_all = {
        key: sum((m.get(key) or 0) for _, m in columns)
        for _, _, _, key, _, _, _, _ in ROWS_DEF if key
    }

    for ri, (_, stt, label, key, bold, is_section, is_sub, note) in enumerate(ROWS_DEF, start=5):
        note_text = note(luy_ke_all, kh_active) if callable(note) else (note or None)
        ws.row_dimensions[ri].height = 40 if note_text else 20

        if is_section:
            bg = C_SECTION
        elif bold and key is not None:
            bg = C_BOLD_ROW
        elif is_sub:
            bg = C_SUBITEM
        else:
            bg = C_NORMAL

        stt_cell = ws.cell(ri, 1, stt)
        stt_cell.border = _border(); stt_cell.alignment = _align()
        if is_section and stt:
            stt_cell.fill = _fill(C_SECTION_NO); stt_cell.font = _font(bold=True, color="FFFFFF")
        else:
            stt_cell.fill = _fill(bg); stt_cell.font = _font(bold=bold)

        lbl_cell = ws.cell(ri, 2, label)
        lbl_cell.fill = _fill(bg); lbl_cell.font = _font(bold=bold, color="1F3864" if is_section else "000000")
        lbl_cell.alignment = _align(h="left"); lbl_cell.border = _border()

        day_vals = []
        for ci, (_, m) in enumerate(columns, start=3):
            val = m.get(key) if key else None
            cell = ws.cell(ri, ci, val)
            cell.fill = _fill(bg); cell.font = _font(bold=bold)
            cell.alignment = _align(); cell.border = _border()
            if val is not None:
                day_vals.append(val)

        luy_ke_val = sum(day_vals) if key and day_vals else None
        lk_cell = ws.cell(ri, luy_ke_col, luy_ke_val)
        lk_cell.fill = _fill(C_LUY_KE if luy_ke_val is not None else bg)
        lk_cell.font = _font(bold=bold); lk_cell.alignment = _align(); lk_cell.border = _border()

        ty_le_val = luy_ke_val / kh_active if (luy_ke_val is not None and kh_active) else None
        tl_cell = ws.cell(ri, ty_le_col, ty_le_val)
        tl_cell.fill = _fill(C_TY_LE if ty_le_val is not None else bg)
        tl_cell.font = _font(bold=bold); tl_cell.alignment = _align(); tl_cell.border = _border()
        if ty_le_val is not None:
            tl_cell.number_format = "0.00%"

        note_cell = ws.cell(ri, note_col, note_text)
        note_cell.fill = _fill(bg); note_cell.font = _font(size=10)
        note_cell.alignment = _align(h="left", wrap=True); note_cell.border = _border()

    ws.freeze_panes = "C5"


def build_report_mb(
    products: dict[str, list[dict]],
    kh_active_by_product: dict[str, int],
) -> bytes:
    """Build bc_cskh_mb (reclassified) report. Returns Excel bytes."""
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, data in products.items():
        ws = wb.active if first else wb.create_sheet(sheet_name)
        if first:
            ws.title = sheet_name
            first = False
        kha = kh_active_by_product.get(sheet_name, 0)
        _populate_sheet_mb(ws, data, kha, product_name=sheet_name)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
