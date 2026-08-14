"""Build the 'thật' CSKH report — returns bytes (BytesIO), no disk I/O."""
from __future__ import annotations

import io
import sys
from collections import defaultdict
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from pipeline.calc import (
    calc_real, VALID_DATE_RANGE, HARDCODED_METRICS_REAL,
    PRODUCT_START_DATES, _build_columns,
)

# ── Colours ───────────────────────────────────────────────────────────────────
C_TITLE      = "1F3864"
C_SUBTITLE   = "2E75B6"
C_HEADER     = "1F4E79"
C_SECTION_NO = "1F4E79"
C_SECTION    = "BDD7EE"
C_BOLD_ROW   = "DEEAF1"
C_NORMAL     = "FFFFFF"
C_SUBITEM    = "EBF3FB"
C_LUY_KE     = "E2EFDA"
C_TY_LE      = "FCE4D6"
C_BIEN_DONG  = "FFF2CC"


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=11, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)


def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border():
    thin = Side(style="thin", color="B0B0B0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _fmt_day(d: date) -> str:
    return d.strftime("%-d/%m/%Y") if sys.platform != "win32" else d.strftime("%#d/%m/%Y")


ROWS_DEF = [
    (0, "",  "Tổng Số lượng cuộc gọi",                           "total",                  True,  False, False),
    (0, "1", "Thái độ khách hàng",                               None,                     True,  True,  False),
    (0, "",  "Hài lòng",                                         "thai_do_hai_long",        False, False, False),
    (0, "",  "Bình thường",                                      "thai_do_binh_thuong",     False, False, False),
    (0, "",  "Gay gắt",                                          "thai_do_gay_gat",         False, False, False),
    (0, "2", "Kênh tiếp nhận",                                   None,                     True,  True,  False),
    (0, "",  "Kênh bán",                                         "mb_total",               True,  False, False),
    (0, "",  "Call",                                              "mb_call",                False, False, False),
    (0, "",  "Email",                                             "mb_email",               False, False, False),
    (0, "",  "Fanpage / mạng xã hội / website",                  "mb_fanpage",             False, False, False),
    (0, "",  "CTBH",                                             "ctbh_total",             True,  False, False),
    (0, "",  "Call",                                             "ctbh_call",              False, False, False),
    (0, "",  "Email",                                            "ctbh_email",             False, False, False),
    (0, "",  "Fanpage / mạng xã hội / Excel",                   "ctbh_fanpage",           False, False, False),
    (0, "",  "LiteX",                                             "litex_total",            True,  False, False),
    (0, "",  "Số cuộc gọi vào",                                   "litex_goi_vao",          False, False, False),
    (0, "",  "Số cuộc gọi ra",                                    "litex_goi_ra",           False, False, False),
    (1, "",  "- Số cuộc gọi KH không trả lời",                   "litex_kh_ko_tl",         False, False, True),
    (0, "",  "Email",                                             "litex_email",            False, False, False),
    (0, "",  "Fanpage / mạng xã hội / website",                  "litex_fanpage",          False, False, False),
    (0, "3", "Phân loại cuộc gọi",                               None,                     True,  True,  False),
    (0, "",  "Khách hàng hỏi quyền lợi sản phẩm / thao tác",    "tu_van",                 False, False, False),
    (0, "",  "Khách hàng yêu cầu hủy dịch vụ",                  "huy",                    False, False, False),
    (1, "",  "- Số khách hàng hủy",                              "so_huy",                 False, False, True),
    (1, "",  "- Số khách hàng tiếp tục sử dụng",                 "so_tiep_tuc",            False, False, True),
    (1, "",  "- Không có kết quả",                               "so_chua_kq",             False, False, True),
    (2, "",  "  + KH không nghe máy / ngắt kết nối",            "so_chua_kq_ko_nghe",     False, False, True),
    (2, "",  "  + KH đã được hỗ trợ / không cần hỗ trợ",        "so_chua_kq_da_ho_tro",  False, False, True),
    (2, "",  "  + Chưa đủ thông tin",                            "so_chua_kq_tt",          False, False, True),
    (2, "",  "  + Khác",                                         "so_chua_kq_khac",        False, False, True),
    (0, "",  "Khách hàng có các khiếu nại khác",                 "khac",                   False, False, False),
    (1, "",  "- KH không nghe máy / ngắt kết nối",              "khac_ko_nghe",           False, False, True),
    (1, "",  "- KH đã được hỗ trợ / không cần hỗ trợ",          "khac_da_ho_tro",         False, False, True),
    (1, "",  "- KH gọi nhầm tổng đài",                          "khac_goi_nham",          False, False, True),
    (1, "",  "- Khác",                                           "khac_khac",              False, False, True),
    (0, "",  "Khách hàng yêu cầu bồi thương",                   "boi_thuong",             False, False, False),
    (0, "",  "Khách hàng khiếu nại lỗi thu phí",               "loi_thu_phi",            False, False, False),
]


def _get_date(r: dict):
    return r.get("date") or r.get("event_date")


def _populate_sheet_real(ws, data: list[dict], kh_active: int, product_name: str = ""):
    data = [r for r in data if VALID_DATE_RANGE[0] <= _get_date(r) <= VALID_DATE_RANGE[1]]

    product_start = PRODUCT_START_DATES.get(product_name)
    if product_start:
        data = [r for r in data if _get_date(r) >= product_start]

    dates = sorted({_get_date(r) for r in data})
    by_date: dict[date, list] = defaultdict(list)
    for r in data:
        by_date[_get_date(r)].append(r)

    seen_bt: set = set()
    metrics_by_date = {d: calc_real(by_date[d], seen_bt) for d in dates}
    if product_name == "Mất Tiền MB Đền":
        metrics_by_date.update(HARDCODED_METRICS_REAL)
        all_dates = sorted(set(dates) | set(HARDCODED_METRICS_REAL))
    else:
        all_dates = sorted(set(dates))

    columns = _build_columns(all_dates, metrics_by_date)

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    for i in range(len(columns)):
        ws.column_dimensions[get_column_letter(i + 3)].width = 14
    luy_ke_col    = len(columns) + 3
    ty_le_col     = luy_ke_col + 1
    bien_dong_col = ty_le_col + 1
    ws.column_dimensions[get_column_letter(luy_ke_col)].width    = 14
    ws.column_dimensions[get_column_letter(ty_le_col)].width     = 22
    ws.column_dimensions[get_column_letter(bien_dong_col)].width = 22
    total_cols = bien_dong_col

    date_range_str = (
        f"{all_dates[0].strftime('%d/%m/%Y')} - {all_dates[-1].strftime('%d/%m/%Y')}"
        if len(all_dates) > 1 else all_dates[0].strftime("%d/%m/%Y")
    )
    timestamp_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    def merge_and_style(row, c1, c2, text, bg, fg="FFFFFF", bold=True, size=13, italic=False, h="center"):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row, c1, text)
        cell.fill = _fill(bg); cell.font = _font(bold=bold, color=fg, size=size, italic=italic)
        cell.alignment = _align(h=h); cell.border = _border()

    ws.row_dimensions[1].height = 36
    title = f"Báo cáo dịch vụ khách hàng - Sản phẩm {product_name}" if product_name else "Báo cáo dịch vụ khách hàng - MB"
    merge_and_style(1, 1, total_cols, title, C_TITLE, size=16)

    ws.row_dimensions[2].height = 20
    merge_and_style(2, 1, total_cols, f"Thời gian lấy dữ liệu: {date_range_str}", C_SUBTITLE, size=12)

    ws.row_dimensions[3].height = 20
    merge_and_style(3, 1, total_cols,
                    f"Số khách hàng active: {kh_active:,}    |    Cập nhật lần cuối: {timestamp_str}",
                    C_SUBTITLE, size=12, bold=False, italic=True)

    ws.row_dimensions[4].height = 30
    headers = ["STT", "Chỉ tiêu / Nội dung"] + [lbl for lbl, _ in columns] + ["Lũy kế", "Tỷ lệ so với số KH active", "Biến động so với ngày trước"]
    for c, hdr in enumerate(headers, start=1):
        cell = ws.cell(4, c, hdr)
        cell.fill = _fill(C_HEADER); cell.font = _font(bold=True, color="FFFFFF", size=11)
        cell.alignment = _align(); cell.border = _border()

    luy_ke_all = {
        key: sum((m.get(key) or 0) for _, m in columns)
        for _, _, _, key, *_ in ROWS_DEF if key
    }
    last_m = columns[-1][1] if columns else {}
    prev_m = columns[-2][1] if len(columns) >= 2 else {}

    for ri, (_, stt, label, key, bold, is_section, is_sub) in enumerate(ROWS_DEF, start=5):
        ws.row_dimensions[ri].height = 20
        if is_section:
            bg = C_SECTION
        elif bold and key is not None:
            bg = C_BOLD_ROW
        elif is_sub:
            bg = C_SUBITEM
        else:
            bg = C_NORMAL

        stt_cell = ws.cell(ri, 1, stt)
        stt_cell.border = _border(); stt_cell.alignment = _align()
        if is_section and stt:
            stt_cell.fill = _fill(C_SECTION_NO); stt_cell.font = _font(bold=True, color="FFFFFF")
        else:
            stt_cell.fill = _fill(bg); stt_cell.font = _font(bold=bold)

        lbl_cell = ws.cell(ri, 2, label)
        lbl_cell.fill = _fill(bg); lbl_cell.font = _font(bold=bold, color="1F3864" if is_section else "000000")
        lbl_cell.alignment = _align(h="left"); lbl_cell.border = _border()

        day_vals = []
        for ci, (_, m) in enumerate(columns, start=3):
            val = m.get(key) if key else None
            cell = ws.cell(ri, ci, val)
            cell.fill = _fill(bg); cell.font = _font(bold=bold)
            cell.alignment = _align(); cell.border = _border()
            if val is not None:
                day_vals.append(val)

        luy_ke_val = sum(day_vals) if key and day_vals else None
        lk_cell = ws.cell(ri, luy_ke_col, luy_ke_val)
        lk_cell.fill = _fill(C_LUY_KE if luy_ke_val is not None else bg)
        lk_cell.font = _font(bold=bold); lk_cell.alignment = _align(); lk_cell.border = _border()

        ty_le_val = luy_ke_val / kh_active if (luy_ke_val is not None and kh_active) else None
        tl_cell = ws.cell(ri, ty_le_col, ty_le_val)
        tl_cell.fill = _fill(C_TY_LE if ty_le_val is not None else bg)
        tl_cell.font = _font(bold=bold); tl_cell.alignment = _align(); tl_cell.border = _border()
        if ty_le_val is not None:
            tl_cell.number_format = "0.00%"

        bd_val = ((last_m.get(key) or 0) - (prev_m.get(key) or 0)) / kh_active \
            if (key and prev_m and kh_active) else None
        bd_cell = ws.cell(ri, bien_dong_col, bd_val)
        bd_cell.fill = _fill(C_BIEN_DONG if bd_val is not None else bg)
        bd_cell.font = _font(bold=bold); bd_cell.alignment = _align(); bd_cell.border = _border()
        if bd_val is not None:
            bd_cell.number_format = "0.00%"

    if all_dates and kh_active:
        total_lk    = luy_ke_all.get("total", 0)
        gay_gat_lk  = luy_ke_all.get("thai_do_gay_gat", 0)
        huy_lk      = luy_ke_all.get("huy", 0)
        tiep_tuc_lk = luy_ke_all.get("so_tiep_tuc", 0)
        ty_le_giu_chan = f"{tiep_tuc_lk / huy_lk:.2%}" if huy_lk else "N/A"
        note_text = (
            f"• Thắc mắc/khiếu nại về sản phẩm đến hết ngày "
            f"{all_dates[-1].strftime('%d/%m/%Y')} là {total_lk:,} trên {kh_active:,} "
            f"khách hàng đăng ký thành công.\n"
            f"• Tỷ lệ khách hàng phản ứng gay gắt là {gay_gat_lk / kh_active:.2%}.\n"
            f"• Tỷ lệ thắc mắc/khiếu nại so với tổng số KH là "
            f"{total_lk / kh_active:.2%} so với trung bình trên các kênh khác là khoảng 1%.\n"
            f"• Tỷ lệ giữ chân khách hàng tại tổng đài LiteX là {ty_le_giu_chan}.\n"
            f"• Nếu trừ đi lượng KH tiếp tục sử dụng thì tỷ lệ thắc mắc/khiếu nại "
            f"so với tổng số KH là {(total_lk - tiep_tuc_lk) / kh_active:.2%}."
        )
        note_row = 5 + len(ROWS_DEF) + 1
        ws.row_dimensions[note_row].height = 100
        ws.merge_cells(
            start_row=note_row, start_column=1,
            end_row=note_row, end_column=total_cols,
        )
        nc = ws.cell(note_row, 1, note_text)
        nc.fill = _fill("FFF9C4")
        nc.font = _font(size=10)
        nc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        nc.border = _border()

    ws.freeze_panes = "C5"


def build_report_real(
    products: dict[str, list[dict]],
    kh_active_by_product: dict[str, int],
) -> bytes:
    """Build bc_cskh (thật) report. Returns Excel bytes."""
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, data in products.items():
        ws = wb.active if first else wb.create_sheet(sheet_name)
        if first:
            ws.title = sheet_name
            first = False
        kha = kh_active_by_product.get(sheet_name, 0)
        _populate_sheet_real(ws, data, kha, product_name=sheet_name)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
